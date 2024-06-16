import os
import subprocess
import sys
from datetime import timedelta

import numpy as np
import openpyxl as on
import xlsxwriter

import time


def minutes(t):  # откидывает часы оставляя только мину и переводит их в инт
    ms = t.seconds / 3600
    ms = ms % 1
    ms *= 60
    return round(ms)


def translation(t):  # все время переводит в минуты а потом в инт
    ms = t.seconds / 3600
    ms *= 60
    return round(ms)


# изменяем срез если значение было без 0 в начале (если конвертированно в строчку, то она получается без 0)
def protection_time(t):
    if type(t) != str:
        t = str(t)
    if len(t) == 7:  # это жопа, когда строчка получается без 0 в начале (тип 7:00, вместо 07:00, то вся это херня ломается)
        hours = int(t[:1])
        minute = int(t[2:4])
    else:
        hours = int(t[:2])
        minute = int(t[3:5])
    return [hours, minute]


def arithmetic(num1, num2, num_operator):
    op = {'+': lambda number_1, number_2: number_1 + number_2,
          '-': lambda number_1, number_2: number_1 - number_2,
          '*': lambda number_1, number_2: number_1 * number_2,
          '/': lambda number_1, number_2: number_1 / number_2, }
    try:
        return op[num_operator](num1, num2)
    except KeyError:
        return 'неизвестная операция'


# options = sys.argv[:]

name = []  # Массив с именами фалов
name_do_not_change = []  # массив с номерами маршуротов которые нельзя менять
row_max = 0  # Масикамальное кол-во строк по всем таблицам
col_max = 0  # Максимальное кол-во столбцов по всем таблицам
min_DY = 2  # Минимальное кол-во остановок в дублирующем участке

# -------Кадрирование считываемых таблиц--------
shift_y = 6  # Сдвиг по оси y
shift_x = 1  # Сдвиг по оси x
# -------------------------------------------

len_row = []  # Число строк в каждой таблице (массив)
len_col = []  # Число колонок в каждой таблице (массив)

name_city = 'Речица'

name_folder_with_schedule = f'./Расписание других городов/{name_city}/schedule/'  # Название каталога с расписанием
name_folder_with_zeros_schedule = f'./Расписание других городов/{name_city}/zeros_schedule/'  # Название каталога с якорями

content = os.listdir(name_folder_with_schedule)  # Открывем католог для последущего сканирования файла

for i in range(len(content)):  # Запись имен файлов в массив name
    counter = 0
    str1 = ''
    while content[i][counter] != '.':
        str1 += content[i][counter]
        counter += 1
    name.append(str1)
    name_do_not_change.append(1)
    # if name[i]=="А7" or name[i]=="А7в":
    #      name_do_not_change[i]=0

for x in range(len(name)):  # считаем максимальные строки и столбцы
    wb = on.reader.excel.load_workbook(filename=name_folder_with_schedule + name[x] + ".xlsx")
    sheet = wb.active

    len_row.append(sheet.max_row - shift_y)  # Число строк(начинаем считать с 1) для каждой таблицы
    len_col.append(sheet.max_column - shift_x)  # Число столбцов(начинаем считать с 1) для каждой таблицы

    if sheet.max_row > row_max:
        row_max = sheet.max_row
    if sheet.max_column > col_max:
        col_max = sheet.max_column

schedule = np.empty((len(name), row_max + 1, col_max), np.object_)  # Создаем пустой трехмерный массив)
schedule_with_anchor = np.empty((len(name), row_max + 1, col_max), np.object_)
# лист с номерами столбцов начала обратного направления в каждом автобусе по остановкам(начинает счиать колонки с 1)
col_reverse = []

for z in range(len(name)):  # Заполняем массив schedule данными из таблиц
    schedule[z][0][0] = name[z]
    wb = on.reader.excel.load_workbook(filename=name_folder_with_schedule + name[z] + ".xlsx")
    sheet = wb.active

    schedule_with_anchor[z][0][0] = name[z]
    wb_anchor = on.reader.excel.load_workbook(filename=name_folder_with_zeros_schedule + name[z] + ".xlsx")
    sheet_anchor = wb_anchor.active

    ign_col = len_col[z] + 1
    # флаг нужен для того чтобы если не было обратного напрвавления
    flag_reverse = False
    for y in range(1, len_row[z] + 1):  # начинаем с одного так в нуле название таблице
        counter_x = 0  # счетчик по стобцам ка х только без учета нон
        for x in range(1, len_col[z] + 1):

            if y == 1:
                if sheet.cell(row=y + shift_y, column=x + shift_x).value is None:
                    ign_col = x  # если клетка в название остоновок нон
                else:  # заполняем лист col_reverse
                    if sheet.cell(row=shift_y, column=x + shift_x).value == "Обратное направление":
                        col_reverse.append(counter_x)  # указывает на первый столбец обратного направления
                        flag_reverse = True

            if ign_col != x:
                schedule[z][y][counter_x] = sheet.cell(row=y + shift_y, column=x + shift_x).value
                schedule_with_anchor[z][y][counter_x] = sheet_anchor.cell(row=y + shift_y, column=x + shift_x).value
                counter_x += 1

        #  если в таблице не было обратного направления то просто записываем ее конец
        if y == 1:
            if not flag_reverse:
                col_reverse.append(len_col[z])

# for y in range(1,len_row[0]+1):
#     print()
#     for x in range(len_col[0]):
#         print(schedule_zeros[0][y][x], end='   ')


"""
                Навигация по массиву schedule
             1-ый индекс - Имя таблицы (без расширения)
             2-ой индекс - Строки в таблице (Начинаем с 1)
             3-ый индекс - Столбцы в таблице (Начинаем c 0)
             print(schedule[0][1][3]) это будет остановка 
"""
# ---------------------------Считывыем массив с ID---------------------------
wb = on.reader.excel.load_workbook(filename=f'./Расписание других городов/{name_city}/ID.xlsx')
sheet_id = wb.active

quantity_row = sheet_id.max_row
quantity_col = sheet_id.max_column

mas_id = np.empty((quantity_row, quantity_col), np.object_)
for row in range(quantity_row):
    for col in range(quantity_col):
        mas_id[row][col] = sheet_id.cell(row=row + 1, column=col + 1).value
# ---------------------------КОНЕЦ Считывыем массив с ID---------------------------

id_bus_stops = []
name_bus_stops = []

for NAME in range(len(name)):
    id_bus_stops.append([])
    name_bus_stops.append([])
    for row in range(1, quantity_row):  # начинаем с 1 потому то в 0 строке заголовок таблици
        # если номер маршрута совпал в таблице, то просто по порядку записываем индексы.
        # если маршруты и это прямое и обратное направление
        if name[NAME][1:] == str(mas_id[row][0]) and (mas_id[row][4] == 11 or mas_id[row][4] == 10):
            id_bus_stops[NAME].append(mas_id[row][1])  # записываем id в двухмерный массив
            name_bus_stops[NAME].append(mas_id[row][2])

# for i in range(len(name)):
#     print(name[i], " reverse= ", col_reverse[i])
#     print(id_bus_stops[i])
# print("")

temp_mas = []
temp_mas_number = []
temp_mas_name = []

for NAME in range(len(name)):
    flag_temp_dy = True
    temp_mas.append([])
    temp_mas_number.append([])
    temp_mas_name.append([])
    for i in range(len(id_bus_stops[NAME])):
        if i - 1 == len(id_bus_stops[NAME]) - min_DY:
            break
        if flag_temp_dy:
            temp_dy = []  # промежуточная переменная для нахождения ДУ
            temp_dy_number = []
            temp_name = []
        else:
            break

        counter = 0

        for temp in range(i, min_DY + i):
            counter += 1
            if i > len(id_bus_stops[NAME]):
                flag_temp_dy = False
                break

            temp_dy.append(id_bus_stops[NAME][temp])
            temp_name.append(name_bus_stops[NAME][temp])

            # добавляем 2 потому что, i идет отсчет от 0, а col_reverse с учетом 2 двух первых клеток в таблице (которую считываем schedule)
            temp_dy_number.append(i + 2 + counter)
            flag_counter = True
        if len(temp_dy) != 0:
            temp_mas[NAME].append(temp_dy)
            temp_mas_number[NAME].append(temp_dy_number)
            temp_mas_name[NAME].append(temp_name)
# print(temp_mas)
# print(temp_mas_name)
# print(temp_mas_number)
DY = []
DY_stop = []
DY_stop_id = []
DY_stop_number = []
for i1 in range(len(temp_mas)):
    for j1 in range(len(temp_mas[i1])):
        flag = True
        flag_first = True
        for i2 in range(i1 + 1, len(temp_mas)):
            for j2 in range(len(temp_mas[i2])):
                if temp_mas[i1][j1] == temp_mas[i2][j2]:
                    if flag:
                        DY.append([])
                        DY_stop.append([])
                        DY_stop_id.append([])
                        DY_stop_number.append([])
                        flag = False
                    if flag_first:
                        DY[len(DY) - 1].append(schedule[i1][0][0])
                        DY_stop_number[len(DY) - 1].append(temp_mas_number[i1][j1][0])
                        for count in range(len(temp_mas_name[i1][j1])):
                            DY_stop[len(DY) - 1].append(temp_mas_name[i1][j1][count])
                            DY_stop_id[len(DY) - 1].append(temp_mas[i1][j1][count])
                        flag_first = False
                    DY[len(DY) - 1].append(schedule[i2][0][0])
                    DY_stop_number[len(DY) - 1].append(
                        temp_mas_number[i2][j2][0])  # первая остановка являеться базвовой

# --------------------------------------------------------------------------------------------------
# ------------------------------------ПОИСК ДУ(СОВПАДЕНИЕ ПО 2 ОТСАНОВКАМ)--------------------------
# --------------------------------------------------------------------------------------------------
file = open(f'./Расписание других городов/{name_city}/Первоначальные ДУ.txt', 'w')

for i in range(len(DY)):
    print("-", i, DY[i], file=file)
    print(DY_stop[i], file=file)
    print(DY_stop_id[i], file=file)
    print(DY_stop_number[i], file=file)
    print("", file=file)

live_or_death = [1 for i in range(len(DY))]

# фильтрация первая. В ней склеиваются ДУ из просто пар в полноценные, а лишнее умирают
for dy1 in range(len(DY)):
    number = 0
    for dy2 in range(dy1 + 1, len(DY)):
        if DY[dy1] == DY[dy2]:
            if DY_stop_id[dy1][len(DY_stop_id[dy1]) - 1] == DY_stop_id[dy2][0]:
                flag = True
                number += 1
                for i in range(len(DY_stop_number[dy1])):
                    if round(int(DY_stop_number[dy2][i]) - int(DY_stop_number[dy1][i])) != number:
                        flag = False
                        break
                if flag:
                    DY_stop[dy1].append(DY_stop[dy2][len(DY_stop_id[dy2]) - 1])
                    DY_stop_id[dy1].append(DY_stop_id[dy2][len(DY_stop_id[dy2]) - 1])
                live_or_death[dy2] = 'death'  # везде ниже заменил на live_or_death[dy2]

# фильтрация вторая. Если совпали id остановок и потом совпали автобусы с каким нибудь ду то убиваем
for dy1 in range(len(DY)):
    for dy2 in range(dy1 + 1, len(DY)):
        sett = len(set(DY_stop_id[dy1]) & set(DY_stop_id[dy2]))
        if sett == len(DY_stop_id[dy2]) and sett == len(DY_stop_id[dy1]):
            set2 = len(set(DY[dy1]) & set(DY[dy2]))
            if set2 == len(DY[dy2]):
                live_or_death[dy2] = 'death'
            elif set2 == len(DY[dy2]):
                live_or_death[dy1] = 'death'

# фильтрация ТРЕЦИЯ. Ищет в больших ду остановки и дописывает в меньший
for dy1 in range(len(DY)):
    for dy2 in range(dy1 + 1, len(DY)):
        sett = len(set(DY[dy1]) & set(DY[dy2]))
        if len(DY[dy2]) == sett:  # смотрим на общие автобусы, если количество общих автобусов совпало с длинной ВТОРОГО
            # если последняя остановка перовго ду совпала с первой  остановки  второго ду
            if DY_stop_id[dy1][len(DY_stop_id[dy1]) - 1] == DY_stop_id[dy2][0]:
                for plus in range(len(DY_stop[dy1]) - 2, -1, -1):  # то в начала второго добавляем остановки первого
                    DY_stop[dy2].insert(0, DY_stop[dy1][plus])
                    DY_stop_id[dy2].insert(0, DY_stop_id[dy1][plus])
            elif DY_stop_id[dy2][len(DY_stop_id[dy2]) - 1] == DY_stop_id[dy1][0]:
                for plus in range(1, len(DY_stop[dy1])):
                    DY_stop[dy2].append(DY_stop[dy1][plus])
                    DY_stop_id[dy2].append(DY_stop_id[dy1][plus])

        # смотрим на общие автобусы, если количество общих автобусов совпало с длинной ПЕРВОГО
        elif len(DY[dy1]) == sett:
            if DY_stop_id[dy1][len(DY_stop_id[dy1]) - 1] == DY_stop_id[dy2][0]:
                for plus in range(1, len(DY_stop[dy2])):
                    DY_stop[dy1].append(DY_stop[dy2][plus])
                    DY_stop_id[dy1].append(DY_stop_id[dy2][plus])
            elif DY_stop_id[dy2][len(DY_stop_id[dy2]) - 1] == DY_stop_id[dy1][0]:
                for plus in range(len(DY_stop[dy2]) - 2, -1, -1):
                    DY_stop[dy1].insert(0, DY_stop[dy2][plus])
                    DY_stop_id[dy1].insert(0, DY_stop_id[dy2][plus])

# фильтр ЧЕТВЕРТЫЙ. ищет в одинковых остановки и дописывает их с любого места, а лишнее убивает
for dy1 in range(len(DY)):
    for dy2 in range(dy1 + 1, len(DY)):
        if DY[dy1] == DY[dy2]:
            if live_or_death[dy1] != 'death' and live_or_death[dy2] != 'death':
                for run1 in range(len(DY_stop_id[dy1])):
                    if DY_stop_id[dy1][run1] == DY_stop_id[dy2][0]:
                        flag = False
                        for plus in range(len(DY_stop_id[dy2])):
                            if DY_stop_id[dy2][plus] == DY_stop_id[dy1][len(DY_stop_id[dy1]) - 1]:
                                flag = True
                                continue
                            if flag:
                                DY_stop[dy1].append(DY_stop[dy2][plus])
                                DY_stop_id[dy1].append(DY_stop_id[dy2][plus])
                        if flag:
                            live_or_death[dy2] = 'death'

# фильтр ПЯТЫЙ. убивает меньший одинаковый ДУ
for dy1 in range(len(DY)):
    for dy2 in range(dy1 + 1, len(DY)):
        if live_or_death[dy1] == 1:
            if live_or_death[dy2] == 1:
                if DY_stop_id[dy1] == DY_stop_id[dy2]:
                    sett = len(set(DY[dy1]) & set(DY[dy2]))
                    if sett == len(DY[dy2]):
                        live_or_death[dy2] = 'death'

# counter = 0
# for i in range(len(DY)):
#     if live_or_death[i] != 'death':
#         print("-", counter, DY[i])
#         print(DY_stop[i])
#         print(DY_stop_id[i])
#         print(DY_stop_number[i])
#         print()
#         counter += 1

# Исправляем индекс базовой остановки
for i in range(len(DY)):  # ходим по дублирующим участком
    if live_or_death[i] != 'death':  # если ду живой
        for ii in range(len(DY[i])):  # ходим по маршрутам в ду
            for NAME in range(len(name)):  # ходим по всем расписаниям автобусов
                if DY[i][ii] == name[NAME]:  # если маршрут в ду совпал в названии таблицы с расписанием
                    # ходим по индексам остановок в данной таблицы с расписанием
                    for stops in range(len(id_bus_stops[NAME])):
                        # если свопал мндекс остановки в расписании с индексом первой остановки в ду
                        if id_bus_stops[NAME][stops] == DY_stop_id[i][0]:
                            # ходим опять по индексам остановок данного расписания
                            for run in range(len(id_bus_stops[NAME])):
                                # если первый индекс остановки сопвпал и с индексом из расписания
                                if DY_stop_id[i][0] == id_bus_stops[NAME][run]:
                                    # добаввляет два потому что первые две колонки без врменеи (начало рейса и карточка)
                                    DY_stop_number[i][ii] = run + 2
                                    break

Final_DY = []
Final_DY_stop_id = []
counter = 0
for i in range(len(DY_stop_number)):
    if live_or_death[i] != 'death':
        Final_DY.append([DY_stop[i]])
        Final_DY[counter].append(DY[i])
        Final_DY[counter].append(DY_stop_number[i])
        Final_DY_stop_id.append(DY_stop_id[i])
        counter += 1

file = open(f'./Расписание других городов/{name_city}/Поэтапныйреузльтат.txt', 'w')
print("ЭТАП ПЕРВЫЙ", file=file)
print("Выделенные дублирующие участки:\n", file=file)
for i in range(len(Final_DY)):
    print("-", i, "Остановки в ДУ", Final_DY[i][0], file=file)
    print("Автобусы в ДУ:", Final_DY[i][1], file=file)
    print("Индекс базовой остановки:", Final_DY[i][2], file=file)
    print("Id остановак:", Final_DY_stop_id[i], file=file)
    print("", file=file)

print('', file=file)
print("Разделение ДУ по направлениям ", file=file)

DY_back = []
DY_ahead = []
DY_mixed = []

# делим все ДУ на 3 напраавления (ду автобусы которого все в одном направлении(прям или обратном) или в разных направлениях)

for i in range(len(Final_DY)):  # бегаем по всем ду
    count = 0  # счетчик для проверки всех автобусов в ду
    for j in range(len(Final_DY[i][2])):  # бегаем по индексам базовой остановки
        # если индекс базовой остановки находится в обратном направление
        if Final_DY[i][2][j] >= col_reverse[name.index(Final_DY[i][1][j])]:
            count += 1  # то плюс
        elif Final_DY[i][2][j] < col_reverse[name.index(Final_DY[i][1][j])]:  # и на оборот
            count -= 1
    if count == -len(Final_DY[i][2]):  # если все атобусы в одном прямом направление то
        DY_ahead.append(Final_DY[i])
    elif count == len(Final_DY[i][2]):  # если все атобусы в обратном направление
        DY_back.append(Final_DY[i])
    else:
        DY_mixed.append(Final_DY[i])  # если в разных направлениях

kof_stop = 1  # На это число умножается кол-во автобусных остановок(значимость) относительно кол-ва автобусов
kof_bus = 1

# kof_stop = options[2]  # На это число умножается кол-во автобусных остановок(значимость) относительно кол-ва автобусов
# kof_bus = options[3]  # На это число умножается кол-во маршуртов относительно кол-ва остановок
number_respect_ahead = []  # Список со значимостью каждого ДУ в массиве DY (для последующего сравнения)
number_respect_back = []
number_respect_mixed = []

mas_death_ahead = []  # Три массива для гонки в рейтинги чтобы оставить живых и мертвых
mas_death_back = []
mas_death_mixed = []

for i in DY_ahead:
    number_respect_ahead.append(len(i[1]) * kof_bus + len(i[0]) * kof_stop)
    mas_death_ahead.append("live")

for i in DY_back:
    number_respect_back.append(len(i[1]) * kof_bus + len(i[0]) * kof_stop)
    mas_death_back.append("live")

for i in DY_mixed:
    number_respect_mixed.append(len(i[1]) * kof_bus + len(i[0]) * kof_stop)
    mas_death_mixed.append("live")

# -----------------print-------------------------------------------

# print("Маршурты автобусов которые идут в обратном направление:")
# for i in DY_back:
#     print(i)
# print("Маршурты автобусов которые идут в разных направление:")
# for i in DY_mixed:
#     print(i)


for sorter_number in range(len(DY_ahead)):  # обычная сортировка с синхронизацией ДУ с респектом
    max_num = sorter_number
    for search_max in range(sorter_number, len(DY_ahead)):
        if number_respect_ahead[max_num] < number_respect_ahead[search_max]:
            max_num = search_max

    temp_num = number_respect_ahead[sorter_number]
    number_respect_ahead[sorter_number] = number_respect_ahead[max_num]
    number_respect_ahead[max_num] = temp_num

    temp_num = DY_ahead[sorter_number]
    DY_ahead[sorter_number] = DY_ahead[max_num]
    DY_ahead[max_num] = temp_num

for sorter_number in range(len(DY_back)):  # обычная сортировка с синхронизацией ДУ с респектом
    max_num = sorter_number
    for search_max in range(sorter_number, len(DY_back)):
        if number_respect_back[max_num] < number_respect_back[search_max]:
            max_num = search_max
    temp_num = number_respect_back[sorter_number]
    number_respect_back[sorter_number] = number_respect_back[max_num]
    number_respect_back[max_num] = temp_num

    temp_num = DY_back[sorter_number]
    DY_back[sorter_number] = DY_back[max_num]
    DY_back[max_num] = temp_num

for sorter_number in range(len(DY_mixed)):  # обычная сортировка с синхронизацией ДУ с респектом
    max_num = sorter_number
    for search_max in range(sorter_number, len(DY_mixed)):
        if number_respect_mixed[max_num] < number_respect_mixed[search_max]:
            max_num = search_max
    temp_num = number_respect_mixed[sorter_number]
    number_respect_mixed[sorter_number] = number_respect_mixed[max_num]
    number_respect_mixed[max_num] = temp_num

    temp_num = DY_mixed[sorter_number]
    DY_mixed[sorter_number] = DY_mixed[max_num]
    DY_mixed[max_num] = temp_num

print("Маршурты автобусов которые идут в прямом направление:", file=file)
for i in range(len(DY_ahead)):  # Вычисление значимости каждого ДУ
    print(DY_ahead[i], number_respect_ahead[i], file=file)
print("Маршурты автобусов которые идут в обратном направление:", file=file)
for i in range(len(DY_back)):
    print(DY_back[i], number_respect_back[i], file=file)
print("Маршурты автобусов которые идут в разных направление:", file=file)
for i in range(len(DY_mixed)):
    print(DY_mixed[i], number_respect_mixed[i], file=file)
# -----------------print-------------------------------------------
# print("Маршурты автобусов которые идут в прямом направление:")

# обработка ду в прямом направление ________________________________________________________________________________________________________________
mas_do_not_change_ahead = []
for i in range(len(DY_ahead)):  # Вычисление значимости каждого ДУ
    mas_death_ahead.append("live")  # пока что делаем все ду живыми
    mas_do_not_change_ahead.append([])  # добавляем измерение для маршуртов в ду, тобы понмиать какой конкретно живой
    for j in range(len(DY_ahead[i][1])):  # ходим по маршрутам в ду
        for NAME in range(
                len(name_do_not_change)):  # ходим по количеству таблиц, а тут именно по названием маршуртов которые мы считали
            if DY_ahead[i][1][j] == name[NAME]:  # если название маршурта совпало
                mas_do_not_change_ahead[i].append(
                    name_do_not_change[NAME])  # то делаем каждый отдельный автобус (1) то есть знать живым
# print("До изменений:",mas_do_not_change_ahead)
# print("")
# for i in range(len(DY_ahead)):  # Вычисление значимости каждого ДУ
#     print(DY_ahead[i],mas_do_not_change_ahead[i])
# print("")
for i in range(len(DY_ahead)):
    for j in range(i + 1, len(DY_ahead)):
        if len(set(DY_ahead[i][1]) & set(DY_ahead[j][1])) != 0:  # если нет совпадений
            if number_respect_ahead[i] >= number_respect_ahead[j]:  # если респект больше то убиваем с меньшим
                if mas_death_ahead[i] == 'death':  # если массив с большим рейтингом уже мертывый
                    continue
                # счетчик нужен для того чтобы посичтать, может у данного ду, все атобусы были заблокиравонны и там все нули
                counter = 0
                for dy in range(len(DY_ahead[i][1])):  # ходим по маршрутам
                    if mas_do_not_change_ahead[i][dy] == 1:  # если данный маршрут в ду живой
                        counter += 1
                if counter != 0:  # если ду с большим рейтингом кто-то живой есть то убиваем с меньшим
                    mas_death_ahead[j] = 'death'
                else:  # а если нет живых то убиваем с большим рейтингом
                    mas_death_ahead[i] = 'death'
mas_live_bus_ahead = []  # массив с маршрутами которые выйграли по рейтингу

for i in range(len(DY_ahead)):
    if mas_death_ahead[i] == 'live':  # если ду живой
        for j in range(len(DY_ahead[i][1])):  # ходим по маршрутам в данном ду
            if mas_do_not_change_ahead[i][j] != 0:  # если в данном ду автобус не мертвый
                if DY_ahead[i][1][j] not in mas_live_bus_ahead:  # если данного автобуса нет в списке
                    mas_live_bus_ahead.append(DY_ahead[i][1][j])  # добовляем

# print(mas_live_bus_ahead)
for i in range(len(DY_ahead)):  #
    if mas_death_ahead[i] == 'death':  # если ду мертвый
        for j in range(len(DY_ahead[i][1])):  # ходим по его маршрутам
            if mas_do_not_change_ahead[i][j] != 0:  # если маршрут не был выбран против изменений
                if DY_ahead[i][1][j] not in mas_live_bus_ahead:  # если данного автобуса нет в списке побидителей )
                    mas_do_not_change_ahead[i][j] = 1  # то оставляем его живым
                    mas_live_bus_ahead.append(DY_ahead[i][1][j])  # и добовляем в список победителей
                else:
                    mas_do_not_change_ahead[i][j] = 0  # иначе убиваем его

# print("После изменений:",mas_do_not_change_ahead)
for i in range(len(DY_ahead)):
    DY_ahead[i].append(mas_do_not_change_ahead[i])  # записываем в ДУ массив с живыми и мертвыми автобусами
# ________________________________________________________________________________________________________________


# print(mas_death_ahead)
# обработка ду в обратном направление________________________________________________________________________________________________________________


# для Back все также что и для Ahead
mas_do_not_change_back = []
for i in range(len(DY_back)):  # Вычисление значимости каждого ДУ
    mas_death_back.append("live")
    mas_do_not_change_back.append([])
    for j in range(len(DY_back[i][1])):
        for NAME in range(len(name_do_not_change)):
            if DY_back[i][1][j] == name[NAME]:
                mas_do_not_change_back[i].append(name_do_not_change[NAME])
# print("До изменений:",mas_do_not_change_back)

for i in range(len(DY_back)):
    for j in range(i + 1, len(DY_back)):
        if len(set(DY_back[i][1]) & set(DY_back[j][1])) != 0:  # если нет совпадений
            if number_respect_back[i] >= number_respect_back[j]:  # если респект больше то убиваем с меньшим
                if mas_death_back[i] == 'death':
                    continue
                counter = 0
                for dy in range(len(DY_back[i][1])):
                    if mas_do_not_change_back[i][dy] == 1:
                        counter += 1
                if counter != 0:
                    mas_death_back[j] = 'death'
                else:
                    mas_death_back[i] = 'death'
mas_live_bus_back = []

for i in range(len(DY_back)):
    if mas_death_back[i] == 'live':
        for j in range(len(DY_back[i][1])):
            if mas_do_not_change_back[i][j] != 0:
                if DY_back[i][1][j] not in mas_live_bus_back:
                    mas_live_bus_back.append(DY_back[i][1][j])

# print(mas_live_bus_back)
for i in range(len(DY_back)):
    if mas_death_back[i] == 'death':
        for j in range(len(DY_back[i][1])):
            if mas_do_not_change_back[i][j] != 0:
                if DY_back[i][1][j] not in mas_live_bus_back:
                    mas_do_not_change_back[i][j] = 1
                    mas_live_bus_back.append(DY_back[i][1][j])
                else:
                    mas_do_not_change_back[i][j] = 0

# print("После изменений:",mas_do_not_change_back)
for i in range(len(DY_back)):
    DY_back[i].append(mas_do_not_change_back[i])
# ________________________________________________________________________________________________________________

# обработка ду в смешанных направлениеях_________________________________________________________________________

# делаем пока что все атобусы в ду живыми за исключение тех которых сразу нельзя менять
for i in range(len(DY_mixed)):  # Вычисление значимости каждого ДУ
    mas_do_not_change_mixed = []  # добавляем измерение для маршуртов в ду, тобы понмиать какой конкретно живой
    for j in range(len(DY_mixed[i][1])):  # ходим по маршрутам в ду
        for NAME in range(len(name_do_not_change)):
            if DY_mixed[i][1][j] == name[NAME]:
                if name_do_not_change[NAME] == 0:
                    mas_do_not_change_mixed.append(0)
                    break
                else:
                    mas_do_not_change_mixed.append(1)
                    break
    DY_mixed[i].append(mas_do_not_change_mixed)

for i in range(len(DY_mixed)):  # бегаем по всем ду
    for j in range(len(DY_mixed[i][2])):  # бегаем по индексам базовой остановки
        # если индекс базовой остановки находится в обратном направление
        if DY_mixed[i][2][j] >= col_reverse[name.index(DY_mixed[i][1][j])]:
            if DY_mixed[i][1][j] not in mas_live_bus_back:
                if DY_mixed[i][3][j] != 0:
                    mas_live_bus_back.append(DY_mixed[i][1][j])
            else:
                DY_mixed[i][3][j] = 0
        elif DY_mixed[i][2][j] < col_reverse[name.index(DY_mixed[i][1][j])]:  # и на оборот
            if DY_mixed[i][1][j] not in mas_live_bus_ahead:
                if DY_mixed[i][3][j] != 0:
                    mas_live_bus_ahead.append(DY_mixed[i][1][j])
            else:
                DY_mixed[i][3][j] = 0

# print("Ahead")
# for i in DY_ahead:
#     print(i)
# print("Back")
# for i in DY_back:
#     print(i)
# print("Mixed")
# for i in DY_mixed:
#     print(i)

# ________________________________________________________________________________________________________________


filter_DY = []
counter = 0
for i in DY_ahead:
    for j in range(len(i[3])):
        if i[3][j] == 1:
            filter_DY.append([i[0]])
            filter_DY[counter].append(i[1])
            filter_DY[counter].append(i[2])
            filter_DY[counter].append(i[3])
            counter += 1
            break

for i in DY_back:
    for j in range(len(i[3])):
        if i[3][j] == 1:
            filter_DY.append([i[0]])
            filter_DY[counter].append(i[1])
            filter_DY[counter].append(i[2])
            filter_DY[counter].append(i[3])
            counter += 1
            break

for i in DY_mixed:
    for j in range(len(i[3])):
        if i[3][j] == 1:
            filter_DY.append([i[0]])
            filter_DY[counter].append(i[1])
            filter_DY[counter].append(i[2])
            filter_DY[counter].append(i[3])
            counter += 1
            break

# -------------------------------------------------------------------------------------------------
# ------------------- КОНЕЦ ВТОРОГО ЭТАПА ФИЛЬТРАЦИИ (ОСТАВЛЯЕМ ТОЛЬКО САМЫЕ ЗНАЧИМЫЕ ДУ)--------------------
# --------------------------------------------------------------------------------------------------
print("", file=file)
print("После ранжирования ДУ остались:", file=file)
for i in filter_DY:
    print(i, file=file)
# Начинаем считать свдиги времени
print("", file=file)
print("----Оптимизация----", file=file)

# print(col_reverse, '\n', col_max)

file_test = open(f'./Расписание других городов/{name_city}/quantity.txt', 'w')
print(len(filter_DY), file=file_test)
# print('ДЛинна filter_DY=',len(filter_DY))
file_test.close()

print("----filter_DY----", len(filter_DY))
# это для определение доли() на сам алгоритм переменная никак не влияет
quantity_dy = len(filter_DY)

# print("START")
# for i in filter_DY:
#     print(i)
# print("END")
# print(name)
# filter_DY = filter_DY[:int(options[1])]


time_off = timedelta(minutes=0)  # время перерыва между прям и обратным направлением
add_time_end_day = timedelta(minutes=0)  # дополнительное врем в конце дня
add_time_start_day = timedelta(minutes=0)  # дополнительное врем в конце дня
minutes_the_end_ahead = timedelta(minutes=0)

all_check_prov = [0, 0]  # массив для проверки измененией
quantity = -1
for dy in filter_DY:
    quantity += 1
    for i in range(5):
        # print('')
        print('', file=file)
    print('Номер ДУ= ', quantity, file=file)
    print(dy, file=file)
    # print('Номер ДУ= ', quantity)
    # print(dy)
    anchor_compare_time_bus = []  # Массив с якорями
    compare_time_bus = []  # Массив с временим прибытия сравневыемых автобусов на базовую остановку
    for num_now_bus in range(len(dy[1])):  # Бегаем по кол-ву автобусов в данном ДУ
        num_now_tab = name.index(dy[1][num_now_bus])  # Находим индекс автобуса в списке
        anchor_compare_time_bus.append([])
        compare_time_bus.append([])  # Создаем новое измерения для след автобуса
        # Бегаем по строкам времени приходом автобуса на базовую остановку
        for y in range(2, len_row[num_now_tab] + 1):
            temp_compare = schedule[num_now_tab][y][dy[2][num_now_bus]]
            anchor_temp_compare = schedule_with_anchor[num_now_tab][y][dy[2][num_now_bus]]
            if temp_compare is not None:
                # записываем якоря на базовой остановке
                anchor_compare_time_bus[num_now_bus].append(anchor_temp_compare)
                # Записываеться время прибытия на базувую остановку
                compare_time_bus[num_now_bus].append(str(temp_compare))
    time_last_previous_bus = []  # массив для сохранения последнего времени в каждом ду
    time_all_previous_bus = []  # массив для сохранения всего времени ду
    name_previous_bus = []  # и их маршруты
    flag_for_first_bus = True

    flag = True
    flag_first_check = True  # Флаг нужен для первого подсчета отпимального отклонения
    flag_counter = 0  # флаг для посчета нового оптимального врмене

    # Меняем базовый чаc
    # 4, 5, 6, 7, 8,
    for hour in (4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 0, 1):
        flag_counter += 1
        counter = 0  # Счетик кол-во приходов всех автобусов на базовую остановку за базовый час (hour)
        anchor_time_bus_one_hour = []  # Массив с якорям на БО
        anchor_old_time_bus_one_hour = []  # ТАкой же массив, только этот нужен для отката
        time_bus_one_hour = []  # Массив с приходами всех автобусов на базовую остановку за базовый час (hour)
        old_time_bus_one_hour = []  # то же самое что time_bus_one_hour, только его не будем изменять, он нужен для отката назад если что
        index_time_bus_one_hour = []  # Массив с названиями всех автобусов на базовую остановку за базовый час (hour)
        mas_live_bus = []  # Масив с индексами живых или мертвых автобусов приходящих на бащовую остановку
        for y in range(len(compare_time_bus)):  # Бегаем по строкам времения прибытия
            for x in range(len(compare_time_bus[y])):  # Бегаем по столбцам времени прибытия
                if compare_time_bus[y][x] is not None:  # Защита от None
                    # compare_time_bus[y][x] = str(compare_time_bus[y][x])  # защита от нестандартных параметров из экселя, например если эксель сичтает это временем
                    # print(compare_time_bus[y][x], dy[1][y])
                    # Если данный час совпадает с текущим(hour) часом; [:2]- часы, [3:5] - минуты
                    if protection_time(compare_time_bus[y][x])[0] == hour:
                        # Заполняем массив временем приходов всех автобусов за текущий(hour) час
                        time_bus_one_hour.append(compare_time_bus[y][x])
                        # Заполняем массив временем приходов всех автобусов за текущий(hour) час, такой же массив как и верхний, но этот мы не изменяем дальше
                        old_time_bus_one_hour.append(compare_time_bus[y][x])
                        # название маршрутов приходящих на базовую остановку в ду за час
                        index_time_bus_one_hour.append(dy[1][y])
                        # Записываем якоря времени в массив такой же как time_bus
                        anchor_time_bus_one_hour.append(anchor_compare_time_bus[y][x])
                        anchor_old_time_bus_one_hour.append(anchor_compare_time_bus[y][x])
                        mas_live_bus.append(dy[3][y])  # и его индекс живучести
                        counter += 1  # То прибовляем к счетчик
        if counter == 1:  # Если прибытий за данный час меньше 2, то тут нехуй оптимизировать
            # записывакм последнее время в ДУ
            time_last_previous_bus.append(time_bus_one_hour[len(time_bus_one_hour) - 1])
            time_all_previous_bus.append(time_bus_one_hour)  # записывакм все время ДУ
            name_previous_bus.append(index_time_bus_one_hour)
            flag_first_check = False
            continue
        elif counter == 0:
            continue

        translated_time_bus_one_hour = []
        # Сортировка массива time_bus_one_hour от минимального (с синхронизацией массива index_time_bus_one_hour)
        for i in range(len(time_bus_one_hour)):
            translated_time_bus_one_hour.append(translation(
                timedelta(hours=protection_time(time_bus_one_hour[i])[0],
                          minutes=protection_time(time_bus_one_hour[i])[1])))
        temp_num = -1
        for sorter_number in range(len(time_bus_one_hour)):
            low_num = sorter_number
            for search_min in range(sorter_number, len(time_bus_one_hour)):
                if translated_time_bus_one_hour[low_num] > translated_time_bus_one_hour[search_min]:
                    low_num = search_min

            temp_num = translated_time_bus_one_hour[sorter_number]
            translated_time_bus_one_hour[sorter_number] = translated_time_bus_one_hour[low_num]
            translated_time_bus_one_hour[low_num] = temp_num

            temp_num = time_bus_one_hour[sorter_number]
            time_bus_one_hour[sorter_number] = time_bus_one_hour[low_num]
            time_bus_one_hour[low_num] = temp_num

            old_time_bus_one_hour[sorter_number] = old_time_bus_one_hour[low_num]
            old_time_bus_one_hour[low_num] = temp_num

            temp_num = mas_live_bus[sorter_number]
            mas_live_bus[sorter_number] = mas_live_bus[low_num]
            mas_live_bus[low_num] = temp_num

            temp_num = index_time_bus_one_hour[sorter_number]
            index_time_bus_one_hour[sorter_number] = index_time_bus_one_hour[low_num]
            index_time_bus_one_hour[low_num] = temp_num

            temp_num = anchor_time_bus_one_hour[sorter_number]
            anchor_time_bus_one_hour[sorter_number] = anchor_time_bus_one_hour[low_num]
            anchor_time_bus_one_hour[low_num] = temp_num

        # ---------------------------------------------------Конец сортировки массива time_bus_one_hour---------

        first_check_prov = 0  # переменная для проверки стало ли хуже после сдивга
        # записывакм последнее время в ДУ
        time_last_previous_bus.append(time_bus_one_hour[len(time_bus_one_hour) - 1])
        time_all_previous_bus.append(time_bus_one_hour)  # записывакм все время ДУ
        name_previous_bus.append(index_time_bus_one_hour)
        # print()
        # Новый расчет оптимального времени для меж часа--------------------------------------------------------
        if counter == 2:
            time_pause = 60 / counter  # Считаем оптимальное время перерыва между приходами автобуса за час
        else:
            time_pause = 60 / (counter - 1)
        time_pause = timedelta(minutes=int(time_pause))

        if flag_counter > 1:  # если не первый час оптимизации
            # первое время в прошлом часу
            # кол-во прибытий в прошлом часу
            quantity_arrival = len(time_all_previous_bus[len(time_all_previous_bus) - 2]) - 1
            if quantity_arrival <= 2:
                optimal_time_previous_hours = 30
            else:
                temp_time = time_all_previous_bus[len(time_all_previous_bus) - 2][0]

                temp_time = timedelta(hours=protection_time(temp_time)[0], minutes=protection_time(temp_time)[1])
                optimal_time_previous_hours = round(
                    (60 - minutes(temp_time)) / (len(time_all_previous_bus[len(time_all_previous_bus) - 2]) - 1))
                # print('optimal_time_previous_hours', optimal_time_previous_hours, temp_time,
                # len(time_all_previous_bus[len(time_all_previous_bus) - 2]) - 1)

            optimal_time_now_bus = time_pause
            # print("optimal_time_now_bus", optimal_time_now_bus)
            mean_optimal_time_between_hours = round(
                (optimal_time_previous_hours + translation(optimal_time_now_bus)) / 2)

        # ------------------------------------------THE END расчета нового оптимального интервала--------------------------------

        print("-----------------------------", hour, " ЧАС-----------------------------", file=file)
        print("Время прибытия всех автобусов за час:", time_bus_one_hour, file=file)
        print("Маршрут автобуса этого времени:", index_time_bus_one_hour, file=file)
        print("Маршрут которые можно изменять:", mas_live_bus, file=file)
        print("Оптимальное время для всех маршрутов в ДУ", time_pause, file=file)

        # print("------------------------------------------------------------------------")
        # print("-----------------------------", hour, " ЧАС-----------------------------")
        # print("------------------------------------------------------------------------")
        # print("Время прибытия всех автобусов на БО за час:", time_bus_one_hour, mas_live_bus)
        # print("Якоря времени, на БО за час:", anchor_time_bus_one_hour)
        # print("Индекс автобуса этого времени:", index_time_bus_one_hour)
        # print("Оптимальное время для всех маршрутов в ДУ", time_pause)

        # -------------------------------- Поиск уникальных и количество их приходов--------------------------------------------------------------------------
        unique = list(set(index_time_bus_one_hour))  # получаем уникальные значения из массва с номерами маршрутов
        number_of_occurrences = []  # массив с количеством вхоождений уникальных автобусов
        # print("Список уникальных:", unique, file=file)
        # print("Список уникальных:", unique)
        for un in range(len(unique)):  # ходим по кол-ву уникальных автобусов
            counter = 0  # счетчик вхождений
            for x in range(len(index_time_bus_one_hour)):  # ходим по названием маршрутов в ду
                # если совпал уникальный и название маршрута из списка то
                if unique[un] == index_time_bus_one_hour[x]:
                    counter += 1
            number_of_occurrences.append(counter)
        # print("Количество приходов:", number_of_occurrences, file=file)
        # print("Количество приходов:", number_of_occurrences)
        # -------------------------------- Конец--------------------------------------------------------------------------

        # -------------------------------- Первый раз считвыем абсолютную разницы времени для одного маршрута-----------------
        mas_first_check_prov_between_same_bus = []  # массив с конечными результатами

        # ходим по массиву с количеством вхождений(такой же длинны что и уникальные)
        # print(name_previous_bus)
        for number in range(len(number_of_occurrences)):
            if number_of_occurrences[number] >= 2:  # если количчетво вхождений больше чем 2
                unique_time_bus_one_hour = []  # создаем массив для запоминания время
                # ходим по времени прибытия автобусов в ду
                for time_bus_one_hour_x in range(len(time_bus_one_hour)):
                    # если уникальный номер маршрута совпал с номером маршрута из списка
                    if unique[number] == index_time_bus_one_hour[time_bus_one_hour_x]:
                        # то записываем время
                        unique_time_bus_one_hour.append(time_bus_one_hour[time_bus_one_hour_x])

                if len(unique_time_bus_one_hour) == 2:
                    optimal_time = 60 / len(unique_time_bus_one_hour)  # оптимальное время для одного маршрута на ду
                else:
                    optimal_time = 60 / (len(unique_time_bus_one_hour) - 1)

                check_prov_between_same_bus = 0
                # print("Оптимальное время для одного маршрута в ДУ", optimal_time, file=file)

                if not flag_first_check:
                    for i in range(len(time_all_previous_bus[len(time_all_previous_bus) - 2]) - 1, -1, -1):
                        if unique[number] == name_previous_bus[len(time_all_previous_bus) - 2][i]:
                            unique_time_bus_one_hour.insert(0,
                                                            time_all_previous_bus[len(time_all_previous_bus) - 2][
                                                                i])
                            break
                # print("Время прихода уникального маршрута- ", unique_time_bus_one_hour)
                # print("Время прихода уникального маршрута- ", unique_time_bus_one_hour, file=file)
                for i in range(len(unique_time_bus_one_hour)):  # ходим по времени прибытия уникального маршрута

                    if i != len(unique_time_bus_one_hour) - 1:  # защита от выхода

                        t1 = unique_time_bus_one_hour[i]  # берем первое время
                        ti = unique_time_bus_one_hour[i + 1]  # берем следущие вермя
                        # находим разницу между приходами
                        t_dif = translation(
                            timedelta(hours=protection_time(ti)[0], minutes=protection_time(ti)[1])) - translation(
                            timedelta(hours=protection_time(t1)[0], minutes=protection_time(t1)[1]))
                        # вычисляем абсолютную разницу между оптимальной разницой и разницой
                        dif = abs(optimal_time - t_dif)
                        check_prov_between_same_bus += dif  # получаем сумму отклонений, чем больше сумма тем хуже
                #         print(ti, t1, t_dif, dif)
                # print("----")
                mas_first_check_prov_between_same_bus.append(check_prov_between_same_bus)
            else:
                # print("Время прихода уникального маршрута- ", 0)
                mas_first_check_prov_between_same_bus.append(0)
        print("Первоначальное, абсолютное отклонение одного маршрута в ДУ =", mas_first_check_prov_between_same_bus,
              file=file)
        # print("Первоначальное, абсолютное отклонение одного маршрута в ДУ =", mas_first_check_prov_between_same_bus)
        # -------------------------------- Конец--------------------------------------------------------------------------

        # -------------------------------- Первый подсчет абсолютного отклонения для всех маршрутов в ДУ--------------------------------------------------------------------------
        flag_improve = True
        first_check_prov = 0
        for time_bus_one_hour_x in range(len(time_bus_one_hour)):
            if flag_improve:
                ti = time_bus_one_hour[time_bus_one_hour_x]
                t1 = time_all_previous_bus[len(time_all_previous_bus) - 2][
                    len(time_all_previous_bus[len(time_all_previous_bus) - 2]) - 1]
                # находим разницу между приходами
                # из-за библиотеки приходится чекать и добавлять 24 часа
                # так когда отнимаем от 00:00 получается не правильное время
                if ti[:2] == "00":
                    t_dif = (translation(
                        timedelta(hours=protection_time(ti)[0], minutes=protection_time(ti)[1])) + 1440) - \
                            translation(timedelta(hours=protection_time(t1)[0], minutes=protection_time(t1)[1]))
                else:
                    t_dif = translation(
                        timedelta(hours=protection_time(ti)[0], minutes=protection_time(ti)[1])) - translation(
                        timedelta(hours=protection_time(t1)[0], minutes=protection_time(t1)[1]))
                # вычисляем абсолютную разницу между оптимальной разницой и разницой
                dif = abs(minutes(time_pause) - t_dif)
                if flag_first_check:
                    dif = 0
                first_check_prov += dif  # получаем сумму отклонений, чем больше сумма тем хуже
                flag_improve = False
            # print(ti, t1, t_dif, dif)
            if time_bus_one_hour_x != len(time_bus_one_hour) - 1:  # защита от выхода
                t1 = time_bus_one_hour[time_bus_one_hour_x]  # берем первое время
                ti = time_bus_one_hour[time_bus_one_hour_x + 1]  # берем следущие вермя
                t_dif = translation(
                    timedelta(hours=protection_time(ti)[0], minutes=protection_time(ti)[1])) - translation(
                    timedelta(hours=protection_time(t1)[0],
                              minutes=protection_time(t1)[1]))  # находим разницу между приходами
                dif = abs(minutes(
                    time_pause) - t_dif)  # вычисляем абсолютную разницу между оптимальной разницой и разницой
                # print(ti, t1, t_dif, dif)
                first_check_prov += dif  # получаем сумму отклонений, чем больше сумма тем хуже
        print("Первоначальное, абсолютное отклонение по всем маршрутом в ДУ", first_check_prov, file=file)
        # print("Первоначальное, абсолютное отклонение по всем маршрутом в ДУ", first_check_prov, '\n')
        # print()
        # -------------------------------- Конец--------------------------------------------------------------------------

        # 3 массива для запоминание деталий сдвига
        mas_sign = []  # массив со заком свдига
        mas_direction = []  # массив с направелнием сдвига
        mas_shift = []  # массив с временем сдвига
        # массив для записи времени из time_bus только по порядку, чтобы сразу видить в оптимизации
        # оставшиеся два массива анологичные как и обычные, только в этих сразу все по порядку
        right_now_time_bus_one_hour = []
        right_now_index_time_bus_one_hour = []
        right_now_mas_live_bus = []

        # -------------------------------- Непосредственно сама ОПТИМИЗАЦИЯ--------------------------------------------------------------------------
        for time_bus_one_hour_x in range(len(time_bus_one_hour)):
            # если это не первый час оптимизации
            if flag_counter > 1:
                # учет нового оптимального перерыва для первого времени
                # временно сохраняем предыдущий интревал для последущих маршрутов в часу
                temp_time = time_pause
                if time_bus_one_hour_x == 0:
                    # изменяем для первого пребытия автобуса оптимальный перерыв
                    time_pause = timedelta(minutes=mean_optimal_time_between_hours)
                # для всех последщуих(считаться будет только один раз во второй интерации, после того как первое изменилось))
                elif time_bus_one_hour_x == 1:
                    interval = timedelta(hours=protection_time(time_bus_one_hour[time_bus_one_hour_x - 1])[0],
                                         minutes=protection_time(time_bus_one_hour[time_bus_one_hour_x - 1])[1])
                    interval = 60 - minutes(interval)
                    if len(time_bus_one_hour) > 2:
                        time_pause = timedelta(minutes=(round(interval / (len(time_bus_one_hour) - 1))))
                    else:
                        time_pause = temp_time
            # print("TIME_PAUSE=", time_pause)
            # print('--------------------------------NEW TIME BUS--------------------------------')
            num_bus = index_time_bus_one_hour[time_bus_one_hour_x]
            num_tab = name.index(num_bus)  # Ищем имя первого автобуса в массиве name
            num_bus_in_dy = dy[1].index(num_bus)
            # Бегаем вертикально по строкам времени приходом автобуса на базовую остановку
            for y in range(2, len_row[num_tab] + 1):
                # print("schedule[num_tab][y][dy[2][num_bus_in_dy]]",schedule[num_tab][y][dy[2][num_bus_in_dy]], str(schedule[num_tab][y][dy[2][num_bus_in_dy]]) == str(time_bus_one_hour[time_bus_one_hour_x]), time_bus_one_hour[time_bus_one_hour_x])
                # Если время совпало с временем из таблицы(таким образом находим строчку нужного нам рейса)
                if str(time_bus_one_hour[time_bus_one_hour_x]) == str(schedule[num_tab][y][dy[2][num_bus_in_dy]]):

                    # print("time_bus_one_hour[time_bus_one_hour_x]",time_bus_one_hour[time_bus_one_hour_x])
                    card = schedule[num_tab][y][1]
                    # Ищем ближайшую такую же карточку как и у нашего рейса двигаясь (наш рейс -1)
                    for y_card in range(y + 1, len_row[num_tab] + 1):
                        if y_card != len_row[num_tab]:
                            # Сравнием карту нашего рейса с каждым двигаясь вверх
                            if schedule[num_tab][y][1] == schedule[num_tab][y_card][1]:
                                next_card = y_card
                                break
                        else:
                            # Делаем индекс такой чтобы специально выйти за таблицу Чтобы получить None, дальше этот NOne обрабатывается
                            next_card = len_row[num_tab] + 2

                    print("", file=file)

                    # так же дописываем жизнь маршрута для сортировки
                    right_now_mas_live_bus.append(mas_live_bus[time_bus_one_hour_x])

                    # если маршрут можно двигать если его еще не двигали и если время не заякоренно
                    if mas_live_bus[time_bus_one_hour_x] != 0 and anchor_time_bus_one_hour[time_bus_one_hour_x] == 0:
                        # ___________________________ВЫЧЕСЛЕНИЕ ПЕРЕРЫВОВ____________________________________________
                        # Если базовая остановку ду находится в прямом направении
                        if dy[2][num_bus_in_dy] < col_reverse[num_tab]:
                            # print("Базавая остановка находится в прямом направление - ", dy[0][0], file=file)
                            # print()
                            # print("Базавая остановка находится в прямом направление - ", dy[0][0])
                            # ______________________________________начало расчета перерыва между пярямым и обратном направлением _________________________________
                            flag = "Ahead"

                            # Время начала обратного направления
                            time_after_ahead = schedule[num_tab][y][col_reverse[num_tab]]

                            counter = col_reverse[num_tab]  # счетчик с такоже число что и range
                            if time_after_ahead is None:  # если обратное направление начинается с нон
                                # идем до самого конца пока его не найдем
                                for time_after_x in range(col_reverse[num_tab], len_col[num_tab]):
                                    counter += 1
                                    time_after_ahead = schedule[num_tab][y][time_after_x]
                                    if time_after_ahead is not None:  # если нашли то стоп
                                        break
                                    # если не нашли, то присваем время начала следущего рейса с такой же карточкой
                                    if counter == len_col[num_tab]:
                                        time_after_ahead = schedule[num_tab][next_card][2]
                                    if time_after_ahead is None:  # если это последний рейс то просто, записываем последние время
                                        time_after_ahead = schedule[num_tab][y][
                                            col_reverse[num_tab] - 1]  # + minutes_the_end_ahead
                            # время конца прямого направления
                            time_before_ahead = schedule[num_tab][y][col_reverse[num_tab] - 1]
                            if time_before_ahead is None:
                                # идем в обратную сторону пока не найдем время
                                for time_before_ahead_x in range(col_reverse[num_tab] - 1, 1, -1):
                                    time_before_ahead = schedule[num_tab][y][time_before_ahead_x]
                                    if time_before_ahead is not None:
                                        break

                            # Время конца прямого направления
                            time_before_ahead = timedelta(hours=protection_time(time_before_ahead)[0],
                                                      minutes=protection_time(time_before_ahead)[1])

                            # Время начала обратного направления
                            # И если нет обратного направления, то просто делаем перерыв, был равен 0 минут
                            if time_after_ahead is not None:
                                time_after_ahead = timedelta(hours=protection_time(time_after_ahead)[0],
                                                             minutes=protection_time(time_after_ahead)[1])
                            else:
                                time_after_ahead = time_before_ahead
                            break_ahead_between_ahead_back = time_after_ahead - time_before_ahead  # перерыв в прямом направление между прямым и обратным направлением
                            # ______________________________________конец расчета перерыва между прямым и обратным направлением _________________________________

                            # ______________________________________начало расчета перерыва между обратным и пярмым направлением _________________________________
                            previous_card = 0
                            for y_card in range(y - 1, -1, -1):
                                # print(schedule[num_tab][y][1], schedule[num_tab][y_card][1])
                                if schedule[num_tab][y][1] == schedule[num_tab][y_card][1]:
                                    previous_card = y_card
                                    break
                            if previous_card == 0:
                                last_time_back = schedule[num_tab][y][2]
                            else:
                                last_time_back = schedule[num_tab][previous_card][len_col[num_tab]]
                                if last_time_back is None:  # получаем время если автобус не доехал до конечки по расписанию
                                    # идем в обратном направление до времени
                                    for time_x in range(len_col[num_tab], 0, -1):
                                        time = schedule[num_tab][previous_card][time_x]
                                        if time is not None:
                                            last_time_back = time  # Время прибытия автобуса на конечку в обратном направление
                                            break

                            last_time_back = timedelta(hours=protection_time(last_time_back)[0],
                                                       minutes=protection_time(last_time_back)[1])
                            # Первое время прибытия автобуса в прямом направ с той же карточкой что и прыдущего А
                            first_time_ahead = schedule[num_tab][y][0]

                            if first_time_ahead is not None:  # переводим во время
                                first_time_ahead = timedelta(hours=int(first_time_ahead[:2]),
                                                             minutes=int(first_time_ahead[3:5]))

                            if first_time_ahead is None:  # Если последняя карточка то просто считаем что перерыва нет, но можно добавить в конце смены через add_time_end_day
                                first_time_ahead = last_time_back + add_time_end_day

                            break_ahead_between_back_ahead = first_time_ahead - last_time_back

                            #  если получается длинна больше
                            #  это получается из-за библиотеки
                            #  -1 day, 23:57:00 такая херня получается
                            if len(str(break_ahead_between_back_ahead)) > 9:
                                # просто вырезаем надпись с day
                                if str(break_ahead_between_back_ahead)[:1] == '-':
                                    break_ahead_between_back_ahead = str(break_ahead_between_back_ahead)[8:]
                                elif str(break_ahead_between_back_ahead)[:1] == "1":
                                    break_ahead_between_back_ahead = str(break_ahead_between_back_ahead)[7:]

                            if y == 2:
                                break_ahead_between_back_ahead = timedelta(hours=int(0)) + add_time_start_day

                            print("Перерыв между ПРЯМЫМ и ОБРАТНЫМ направлением в прямом направление",
                                  break_ahead_between_ahead_back, "Для времени - ",
                                  time_bus_one_hour[time_bus_one_hour_x], index_time_bus_one_hour[time_bus_one_hour_x],
                                  file=file)
                            # print("Перерыв между ПРЯМЫМ и ОБРАТНЫМ направлением в прямом направление",
                            #       break_ahead_between_ahead_back, "Для времени - ",
                            #       time_bus_one_hour[time_bus_one_hour_x], index_time_bus_one_hour[time_bus_one_hour_x])
                            # print("break_ahead_between_ahead_back---", "first_time_ahead=", time_after_ahead, "last_time_back=", time_before_ahead)

                            print("Перерыв между ОБРАТНЫМ и ПРЯМЫМ направлением в прямом направление",
                                  break_ahead_between_back_ahead, "Для времени - ",
                                  time_bus_one_hour[time_bus_one_hour_x], index_time_bus_one_hour[time_bus_one_hour_x],
                                  file=file)
                            # print("Перерыв между ОБРАТНЫМ и ПРЯМЫМ направлением в прямом направление",
                            #       break_ahead_between_back_ahead, "Для времени - ",
                            #       time_bus_one_hour[time_bus_one_hour_x], index_time_bus_one_hour[time_bus_one_hour_x])
                            # print()
                            # ______________________________________конец расчета перерыва между обратным и пярмым направлением _________________________________

                        # Если базовая остановку ду находится в обратном направении
                        elif dy[2][num_bus_in_dy] >= col_reverse[num_tab]:
                            print("Базавая остановка находится в обратном напрваление направление - ", dy[0][0],
                                  file=file)
                            # print()
                            # print("Базавая остановка находится в обратном напрваление направление - ", dy[0][0])
                            flag = "Back"

                            # Время прибытия автобуса на конечку в прямом направление
                            time_before = schedule[num_tab][y][col_reverse[num_tab] - 1]

                            # получаем время если автобус не доехал до конечки по расписанию
                            if schedule[num_tab][y][col_reverse[num_tab] - 1] is None:
                                # идем в обратном направление до времени
                                for time_x in range(col_reverse[num_tab] - 1, 2, -1):
                                    time_before = schedule[num_tab][y][time_x]
                                    if time_before is not None:
                                        break
                            if time_before is None:
                                time_before = schedule[num_tab][y][0]

                            # Время начала обратного направления
                            time_after = schedule[num_tab][y][col_reverse[num_tab]]
                            if time_after is None:
                                for time_x in range(col_reverse[num_tab], len_col[num_tab]):
                                    time_after = schedule[num_tab][y][time_x]
                                    if time_after is not None:
                                        break

                            time_after = timedelta(hours=protection_time(time_after)[0],
                                                   minutes=protection_time(time_after)[1])
                            time_before = timedelta(hours=protection_time(time_before)[0],
                                                    minutes=protection_time(time_before)[1])
                            # print('time_after= ', time_after, 'time_before= ', time_before)
                            break_back_between_ahead_back = time_after - time_before  # перерыв между прямым и обратным направлением

                            # __________________конец расчета между прямого и обратного

                            # _____________________начало расчета между обртаным и прямым
                            if schedule[num_tab][y][len_col[num_tab] - 1] is not None:
                                last_time = schedule[num_tab][y][len_col[num_tab] - 1]
                            else:
                                for time_x in range(len_col[num_tab] - 1, 1,
                                                    -1):  # идем в обратном направление до времени если None
                                    time = schedule[num_tab][y][time_x]
                                    if time is not None:
                                        last_time = time  # Время прибытия автобуса на конечку в прямом направление
                                        break

                            last_time = timedelta(hours=protection_time(last_time)[0],
                                                  minutes=protection_time(last_time)[1])

                            # Первое время прибытия автобуса в прямом направ с той же карточкой с предыдущей
                            first_time = schedule[num_tab][next_card][2]

                            if first_time is not None:  # переводим в дату
                                first_time = timedelta(hours=protection_time(first_time)[0],
                                                       minutes=protection_time(first_time)[1])

                            if first_time is None:  # Если последняя карточка то просто считаем что перерыва нет, но можно добавить в конце смены через add_time_end_day
                                first_time = last_time + add_time_end_day

                            break_back_between_back_ahead = first_time - last_time

                            print("Перерыв между ПРЯМЫМ и ОБРАТНЫМ ", break_back_between_ahead_back,
                                  "Для времени - ", time_bus_one_hour[time_bus_one_hour_x],
                                  index_time_bus_one_hour[time_bus_one_hour_x], file=file)
                            # print("Перерыв между ПРЯМЫМ и ОБРАТНЫМ ", break_back_between_ahead_back,
                            #       "Для времени - ", time_bus_one_hour[time_bus_one_hour_x],
                            #       index_time_bus_one_hour[time_bus_one_hour_x])

                            print("Перерыв между ОБРАТНЫМ и ПРЯМЫМ", break_back_between_back_ahead,
                                  "Для времени - ", time_bus_one_hour[time_bus_one_hour_x],
                                  index_time_bus_one_hour[time_bus_one_hour_x], file=file)
                            # print("Перерыв между ОБРАТНЫМ и ПРЯМЫМ", break_back_between_back_ahead,
                            #       "Для времени - ", time_bus_one_hour[time_bus_one_hour_x],
                            #       index_time_bus_one_hour[time_bus_one_hour_x])
                        # ____________________________КОНЕЦ ВЫЧЕСЛЕНИЕ ПЕРЕРЫВОВ____________________________________________

                        time_bus = time_bus_one_hour[time_bus_one_hour_x]
                        # время автобуса приходящего на базавую остановку
                        time_bus = timedelta(hours=protection_time(time_bus)[0],
                                             minutes=protection_time(time_bus)[1])

                        if flag == "Ahead":  # если автобус находится в прямом направление
                            break_left = break_ahead_between_back_ahead  # берем переррыв между обратным и прямым направлением в прямом направление
                            break_right = break_ahead_between_ahead_back  # прямым и обратынм
                        elif flag == "Back":
                            break_left = break_back_between_ahead_back  # аналогично break_back_between_ahead_back
                            break_right = break_back_between_back_ahead

                        # если это первый автобус
                        if time_bus_one_hour_x == 0:
                            now_bus = index_time_bus_one_hour[0]  # первый автобус

                            if flag_for_first_bus:
                                # и отнимаем эту разницу от первого времени, как будто с предыдущим автобусом все хорошо
                                # в первом часу оставляем просто первое время
                                time_previous_bus = timedelta(hours=protection_time(time_bus_one_hour[0])[0],
                                                              minutes=protection_time(time_bus_one_hour[0])[1])
                                # time_previous_bus = timedelta(hours=protection_time(time_bus_one_hour[1])[0], minutes=protection_time(time_bus_one_hour[1])[1]) - time_pause
                                flag_for_first_bus = False
                            else:
                                time_previous_bus = time_last_previous_bus[len(time_last_previous_bus) - 2]

                            # перевод в формат времени из строки
                            time_previous_bus = timedelta(hours=protection_time(time_previous_bus)[0],
                                                          minutes=protection_time(time_previous_bus)[1])

                            first_time_bus = time_bus_one_hour[time_bus_one_hour_x]
                            # первое время автобуса
                            first_time_bus = timedelta(hours=protection_time(first_time_bus)[0],
                                                       minutes=protection_time(first_time_bus)[1])

                            # берем следущий автобус приходящий на эту остановку
                            next_time_bus = time_bus_one_hour[time_bus_one_hour_x + 1]
                            next_time_bus = timedelta(hours=protection_time(next_time_bus)[0],
                                                      minutes=protection_time(next_time_bus)[1])

                            time_between_previous_and_first = first_time_bus - time_previous_bus  # время между первым временем и последним в прошлом часу
                            time_between = next_time_bus - time_bus  # разница между след автобусом и сейчас

                            # если вермя между предыщим автобусов в часу равно оптим перерыву
                            if time_between_previous_and_first == time_pause:
                                mas_sign.append(0)
                                mas_shift.append(0)
                                mas_direction.append(0)
                                right_now_time_bus_one_hour.append(time_bus_one_hour[0])
                                right_now_index_time_bus_one_hour.append(now_bus)
                                continue

                            elif translation(time_between_previous_and_first) > 40:

                                #  или если перерыв между предыдущим автобусов из прошлого часа больше чем 40 минут нет(нет смысла его к нему подтягивать)
                                if time_between > time_pause:  # если перерыв между пред А больше чем оптимальная пауза
                                    flag_side = True  # то флаг присваеваем тру, это значит что можно будет двигать в любую сторону, а не только влево
                                    if time_between > break_right:  # если перерыв между автобусами больше чем возможный сдвиг вправо, то
                                        time_shift = break_right  # то присваеваем его
                                    else:
                                        time_shift = time_between  # если меньше, то присваемваем этот перерыв между автобусами
                                elif time_between < time_pause:  # если перерыв между пред А меньши чем оптимальная пауза
                                    flag_side = False  # то сдвиг только влево
                                    if time_pause > break_left:  # если оптимальный перерыв  больше чем левый перерыв
                                        if break_left > timedelta(minutes=minutes(
                                                time_bus)):  # если левый перерыв больше чем оставшаеся минут в часу у данного автобуса
                                            time_shift = timedelta(minutes=minutes(time_bus))
                                        else:
                                            time_shift = break_left
                                    else:  # если оптимальный перерыв меньше чем левый переррыв
                                        if time_pause > timedelta(minutes=minutes(time_bus)):
                                            time_shift = timedelta(minutes=minutes(time_bus))
                                        else:
                                            time_shift = time_pause

                            # если перерыв между пред А БОЛЬШЕ чем оптимальный перерыв
                            elif time_between_previous_and_first > time_pause and translation(
                                    time_between_previous_and_first) < 40:

                                flag_side = False
                                time_difference = time_between_previous_and_first - time_pause  # сколько есть времени до отимального между предыдущим А из прошлого часа
                                time_shift = time_difference  # присваеваем сдвиг до предыдущего автобуса

                                # если сдвиг больше чем осталось минут у автобуса до предыдущего часа
                                if time_difference > timedelta(minutes=minutes(time_bus)):
                                    time_shift = timedelta(minutes=minutes(time_bus))

                                if time_shift > break_left:  # если сдвиг больше чем перерыв между обратным и прямым
                                    time_shift = break_left

                                time_before_hour = minutes(
                                    timedelta(hours=protection_time(time_bus_one_hour[time_bus_one_hour_x])[0],
                                              minutes=protection_time(time_bus_one_hour[time_bus_one_hour_x])[
                                                  1]))

                                time_before_hour = timedelta(minutes=time_before_hour)
                                # если сдвиг больше чем минут до часа, то оставляем минуты до часа
                                if time_shift > time_before_hour:
                                    time_shift = time_before_hour
                            # если перерыв между пред А МЕНЬШИ чем оптимальный перерыв
                            elif time_between_previous_and_first < time_pause:
                                flag_side = True
                                # вычисляем разницу между идеальным перерывом и настоящим перерывом
                                time_before_optimal_break = time_pause - time_between_previous_and_first
                                # присваемаем правый перерыв
                                time_shift = break_right
                                # если правый перерыв больше чем разница между настоящим перерыв и оптимальным
                                if time_shift > time_before_optimal_break:
                                    # то сдивгаем на эту разницу (3 ду 14 час)
                                    time_shift = time_before_optimal_break

                            if flag_side:
                                sign = "+"
                            else:
                                sign = "-"

                            # schedule[num_tab][y][dy[2][num_bus_in_dy]] = arithmetic(time_bus, time_shift, sign)
                            # Также записываем новое время и в массив с ду
                            temp = arithmetic(time_bus, time_shift, sign)
                            time_bus_one_hour[0] = str(timedelta(hours=protection_time(temp)[0],
                                                                 minutes=protection_time(temp)[1]))
                            # так же записываем это время в такой массив как и time_bus только он сразу сортируется
                            right_now_time_bus_one_hour.append(
                                str(timedelta(hours=protection_time(temp)[0], minutes=protection_time(temp)[1])))
                            right_now_index_time_bus_one_hour.append(now_bus)

                            if flag == "Ahead":
                                # идем от начала таблицы до конца прямого направления
                                for shift in range(2, col_reverse[num_tab]):
                                    if schedule[num_tab][y][shift] is not None:
                                        if type(schedule[num_tab][y][shift]) != str:
                                            schedule[num_tab][y][shift] = str(schedule[num_tab][y][shift])
                                        temp = schedule[num_tab][y][shift]
                                        temp = timedelta(hours=protection_time(temp)[0],
                                                         minutes=protection_time(temp)[1])
                                        schedule[num_tab][y][shift] = str(arithmetic(temp, time_shift, sign))
                                        if len(schedule[num_tab][y][shift]) > 9:
                                            # просто вырезаем надпись с day
                                            if schedule[num_tab][y][shift][:1] == '-':
                                                schedule[num_tab][y][shift] = schedule[num_tab][y][shift][8:]
                                            elif str(schedule[num_tab][y][shift][:1]) == "1":
                                                schedule[num_tab][y][shift] = schedule[num_tab][y][shift][7:]

                            elif flag == "Back":
                                # идем от конца таблице до начала обратного направления
                                for shift in range(len_col[num_tab] - 1, col_reverse[num_tab] - 1, -1):
                                    if schedule[num_tab][y][shift] is not None:
                                        temp = schedule[num_tab][y][shift]
                                        temp = timedelta(hours=protection_time(temp)[0],
                                                         minutes=protection_time(temp)[1])
                                        schedule[num_tab][y][shift] = str(arithmetic(temp, time_shift, sign))
                                        if len(schedule[num_tab][y][shift]) > 9:
                                            # просто вырезаем надпись с day
                                            if schedule[num_tab][y][shift][:1] == '-':
                                                schedule[num_tab][y][shift] = schedule[num_tab][y][shift][8:]
                                            elif str(schedule[num_tab][y][shift][:1]) == "1":
                                                schedule[num_tab][y][shift] = schedule[num_tab][y][shift][7:]
                            mas_sign.append(sign)
                            mas_shift.append(translation(time_shift))
                            mas_direction.append(flag)

                        # для последущих автобусов
                        else:
                            now_bus = index_time_bus_one_hour[time_bus_one_hour_x]  # данный автобус
                            # print(time_bus_one_hour[time_bus_one_hour_x - 1], right_now_time_bus_one_hour[time_bus_one_hour_x-1])
                            previous_time_bus = right_now_time_bus_one_hour[time_bus_one_hour_x - 1]
                            previous_time_bus = timedelta(hours=protection_time(previous_time_bus)[0],
                                                          minutes=protection_time(previous_time_bus)[1])
                            now_difference = time_bus - previous_time_bus

                            #  опять обработка надписси с day
                            if len(str(now_difference)) > 9:
                                now_difference = str(now_difference)
                                if now_difference[:1] == '-':
                                    now_difference = now_difference[8:]
                                elif now_difference[:1] == "1":
                                    now_difference = now_difference[7:]
                                #  к оптимальной паузе добавляем время до нового сдвинутого времени кторое перпрыгнуло данной время
                                now_difference = time_pause + timedelta(minutes=(60 - int(now_difference[3:5])))

                            if now_difference == time_pause:  # если сейчас разница времени равна с оптимальным перерывом
                                mas_sign.append(0)
                                mas_shift.append(0)
                                mas_direction.append(0)
                                right_now_time_bus_one_hour.append(time_bus_one_hour[time_bus_one_hour_x])
                                right_now_index_time_bus_one_hour.append(now_bus)
                                continue
                            elif now_difference > time_pause:  # если сейчас разница больше чем оптимальный перерыв
                                flag_difference = True
                                time_shift = now_difference - time_pause  # то присваеваем разницу между настойщим перерывом и оптимыльным перерывом
                                if time_shift > break_left:
                                    time_shift = break_left

                            elif now_difference < time_pause:  # если сейчас разница МЕНЬШИ чем оптимальный перерыв
                                flag_difference = False
                                time_shift = time_pause - now_difference
                                if time_shift > break_right:
                                    time_shift = break_right

                                # узнаем сколько времени до следущего часа, чтобы при оптимизации не выскочить в следущий час при сдвиге вправа
                                # отнимаем минуты чтобы не выйти в новый час
                                time_before_hour = 60 - minutes(
                                    timedelta(hours=protection_time(time_bus_one_hour[time_bus_one_hour_x])[0],
                                              minutes=protection_time(time_bus_one_hour[time_bus_one_hour_x])[
                                                  1])) - 1

                                time_before_hour = timedelta(minutes=time_before_hour)
                                # если сдвиг больше чем минут до часа, то оставляем минуты до часа
                                if time_shift > time_before_hour:
                                    time_shift = time_before_hour

                            if flag_difference:
                                sign = "-"
                                # print("schedule до сдвига", schedule[num_tab][y][dy[2][num_bus_in_dy]])
                                # schedule[num_tab][y][dy[2][num_bus_in_dy]] = arithmetic(time_bus, time_shift, "-")
                                # schedule[num_tab][y][dy[2][num_bus_in_dy]] = time_bus - time_shift
                                # Также записываем новое время и в массив с ду
                                time_bus_one_hour[time_bus_one_hour_x] = str(arithmetic(time_bus, time_shift, sign))
                                # print("schedule после сдвига", schedule[num_tab][y][dy[2][num_bus_in_dy]])
                                right_now_time_bus_one_hour.append(str(arithmetic(time_bus, time_shift, sign)))

                            else:
                                sign = "+"
                                # print("schedule до сдвига", schedule[num_tab][y][dy[2][num_bus_in_dy]])
                                # schedule[num_tab][y][dy[2][num_bus_in_dy]] = arithmetic(time_bus, time_shift, "+")
                                # Также записываем новое время и в массив с ду
                                time_bus_one_hour[time_bus_one_hour_x] = str(arithmetic(time_bus, time_shift, sign))
                                # print("schedule после сдвига", schedule[num_tab][y][dy[2][num_bus_in_dy]])
                                right_now_time_bus_one_hour.append(str(arithmetic(time_bus, time_shift, sign)))

                            right_now_index_time_bus_one_hour.append(index_time_bus_one_hour[time_bus_one_hour_x])

                            if flag == "Ahead":
                                # идем от начала таблицы до конца прямого направления
                                for shift in range(2, col_reverse[num_tab]):
                                    if schedule[num_tab][y][shift] is not None:
                                        if schedule_with_anchor[num_tab][y][shift] != 1:
                                            if type(schedule[num_tab][y][shift]) != str:
                                                schedule[num_tab][y][shift] = str(schedule[num_tab][y][shift])
                                            temp = schedule[num_tab][y][shift]
                                            # print("schedule до сдвига до оптимизации", schedule[num_tab][y][shift])
                                            temp = timedelta(hours=protection_time(temp)[0],
                                                             minutes=protection_time(temp)[1])
                                            schedule[num_tab][y][shift] = str(arithmetic(temp, time_shift, sign))
                                            # print("schedule после сдвига  после оптимизации ", schedule[num_tab][y][shift],"\n")
                                            #  если получается длинна больше
                                            #  это получается из-за библиотеки
                                            #  -1 day, 23:57:00 такая херня получается
                                            if len(schedule[num_tab][y][shift]) > 9:
                                                # просто вырезаем надпись с day
                                                if schedule[num_tab][y][shift][:1] == '-':
                                                    schedule[num_tab][y][shift] = schedule[num_tab][y][shift][8:]
                                                elif str(schedule[num_tab][y][shift][:1]) == "1":
                                                    schedule[num_tab][y][shift] = schedule[num_tab][y][shift][7:]

                            elif flag == "Back":
                                # идем от конца таблице до начала обратного направления
                                for shift in range(len_col[num_tab] - 1, col_reverse[num_tab] - 1, -1):
                                    if schedule_with_anchor[num_tab][y][shift] != 1:
                                        if schedule[num_tab][y][shift] is not None:
                                            temp = schedule[num_tab][y][shift]
                                            # print("schedule до сдвига до оптимизации", schedule[num_tab][y][shift])
                                            temp = timedelta(hours=protection_time(temp)[0],
                                                             minutes=protection_time(temp)[1])
                                            # print("schedule до сдвига", schedule[num_tab][y][shift],"Полсе сдвига= ", str(arithmetic(temp, time_shift, sign)))
                                            schedule[num_tab][y][shift] = str(arithmetic(temp, time_shift, sign))
                                            # print("schedule после сдвига после оптимизации", schedule[num_tab][y][shift])
                                            #  если получается длинна больше
                                            #  это получается из-за библиотеки
                                            #  -1 day, 23:57:00 такая херня получается
                                            if len(schedule[num_tab][y][shift]) > 9:
                                                # просто вырезаем надпись с day
                                                if schedule[num_tab][y][shift][:1] == '-':
                                                    schedule[num_tab][y][shift] = schedule[num_tab][y][shift][8:]
                                                elif str(schedule[num_tab][y][shift][:1]) == "1":
                                                    schedule[num_tab][y][shift] = schedule[num_tab][y][shift][7:]
                            mas_sign.append(sign)
                            mas_shift.append(translation(time_shift))
                            mas_direction.append(flag)

                    else:
                        mas_sign.append(0)
                        mas_shift.append(0)
                        mas_direction.append(0)

                        # сразу дописываем изменненное время чтобы отсортировать и маршурт
                        right_now_time_bus_one_hour.append(time_bus_one_hour[time_bus_one_hour_x])
                        right_now_index_time_bus_one_hour.append(index_time_bus_one_hour[time_bus_one_hour_x])

                    #  перевод времени для сортировки
                    translated_right_now_time_bus_one_hour = []
                    for i in range(len(right_now_time_bus_one_hour)):
                        translated_right_now_time_bus_one_hour.append(translation(
                            timedelta(hours=protection_time(right_now_time_bus_one_hour[i])[0],
                                      minutes=protection_time(right_now_time_bus_one_hour[i])[1])))
                    # сортировка
                    temp_num = -1
                    temp_index = index_time_bus_one_hour
                    for sorter_number in range(len(translated_right_now_time_bus_one_hour)):
                        low_num = sorter_number
                        for search_min in range(sorter_number, len(translated_right_now_time_bus_one_hour)):
                            if translated_right_now_time_bus_one_hour[low_num] > \
                                    translated_right_now_time_bus_one_hour[search_min]:
                                low_num = search_min

                        temp_num = translated_right_now_time_bus_one_hour[sorter_number]
                        translated_right_now_time_bus_one_hour[sorter_number] = translated_right_now_time_bus_one_hour[
                            low_num]
                        translated_right_now_time_bus_one_hour[low_num] = temp_num

                        temp_num = right_now_time_bus_one_hour[sorter_number]
                        right_now_time_bus_one_hour[sorter_number] = right_now_time_bus_one_hour[low_num]
                        right_now_time_bus_one_hour[low_num] = temp_num

                        temp_num = right_now_index_time_bus_one_hour[sorter_number]
                        right_now_index_time_bus_one_hour[sorter_number] = right_now_index_time_bus_one_hour[low_num]
                        right_now_index_time_bus_one_hour[low_num] = temp_num

                        temp_num = right_now_mas_live_bus[sorter_number]
                        right_now_mas_live_bus[sorter_number] = right_now_mas_live_bus[low_num]
                        right_now_mas_live_bus[low_num] = temp_num

                        temp_num = anchor_time_bus_one_hour[sorter_number]
                        anchor_time_bus_one_hour[sorter_number] = anchor_time_bus_one_hour[low_num]
                        anchor_time_bus_one_hour[low_num] = temp_num

                    # print(right_now_time_bus_one_hour)
                    # print(right_now_index_time_bus_one_hour)
                    # print(right_now_mas_live_bus)
                    # print()
        # -------------------------------- Конец ОПТИМИЗАЦИИ --------------------------------------------------------------------------

        # print("Сдвинутое время, без окончательной проверки, до сортировки: ", time_bus_one_hour)
        # print("Маршрут автобуса этого времени, до сортировкой:", index_time_bus_one_hour)

        # -------------------------------------Сортировка-------------------------------------------------------------------------
        # Еще раз делаем сортировку на всякий случай если время после сдвига перескочила в time_bus_hour
        translated_time_bus_one_hour = []
        for i in range(len(time_bus_one_hour)):
            translated_time_bus_one_hour.append(translation(
                timedelta(hours=protection_time(time_bus_one_hour[i])[0],
                          minutes=protection_time(time_bus_one_hour[i])[1])))

        temp_num = -1
        # сохраняем массив с индексами, чтобы если выполнился комбэк
        # и сортировка вдруг перекинула время, вернуть назад
        temp_index = index_time_bus_one_hour
        for sorter_number in range(len(time_bus_one_hour)):
            low_num = sorter_number
            for search_min in range(sorter_number, len(time_bus_one_hour)):

                if translated_time_bus_one_hour[low_num] > translated_time_bus_one_hour[search_min]:
                    low_num = search_min

            temp_num = translated_time_bus_one_hour[sorter_number]
            translated_time_bus_one_hour[sorter_number] = translated_time_bus_one_hour[low_num]
            translated_time_bus_one_hour[low_num] = temp_num

            temp_num = time_bus_one_hour[sorter_number]
            time_bus_one_hour[sorter_number] = time_bus_one_hour[low_num]
            time_bus_one_hour[low_num] = temp_num

            # temp_num = old_time_bus_one_hour[sorter_number]
            # old_time_bus_one_hour[sorter_number] = old_time_bus_one_hour[low_num]
            # old_time_bus_one_hour[low_num] = temp_num

            temp_num = mas_live_bus[sorter_number]
            mas_live_bus[sorter_number] = mas_live_bus[low_num]
            mas_live_bus[low_num] = temp_num

            temp_num = index_time_bus_one_hour[sorter_number]
            index_time_bus_one_hour[sorter_number] = index_time_bus_one_hour[low_num]
            index_time_bus_one_hour[low_num] = temp_num

            temp_num = anchor_time_bus_one_hour[sorter_number]
            anchor_time_bus_one_hour[sorter_number] = anchor_time_bus_one_hour[low_num]
            anchor_time_bus_one_hour[low_num] = temp_num

        print("Сдвинутое время, без окончательной проверки: ", time_bus_one_hour, file=file)
        # print("Сдвинутое время, без окончательной проверки, после сортировки: ", time_bus_one_hour)
        print("Маршрут автобуса этого времени, с новой сортировкой:", index_time_bus_one_hour, file=file)
        # print("Маршрут автобуса этого времени, с новой сортировкой:", index_time_bus_one_hour)
        # -------------------------------------Конец сортировки-------------------------------------------------------------------------

        # -------------------------------- Второй раз считвыем абсолютную разницы времени для одного маршрута-----------------
        mas_second_check_prov_between_same_bus = []  # массив с конечными результатами
        # ходим по массиву с количеством вхождений(такой же длинны что и уникальные)
        for number in range(len(number_of_occurrences)):
            if number_of_occurrences[number] >= 2:  # если количчетво вхождений больше чем 2
                unique_time_bus_one_hour = []  # создаем массив для запоминания время
                # ходим по времени прибытия автобусов в ду
                for time_bus_one_hour_x in range(len(time_bus_one_hour)):
                    # если уникальный номер маршрута совпал с номером маршрута из списка
                    if unique[number] == index_time_bus_one_hour[time_bus_one_hour_x]:
                        # то записываем время
                        unique_time_bus_one_hour.append(time_bus_one_hour[time_bus_one_hour_x])

                check_prov_between_same_bus = 0
                if not flag_first_check:
                    for i in range(len(time_all_previous_bus[len(time_all_previous_bus) - 2]) - 1, -1, -1):

                        if unique[number] == name_previous_bus[len(time_all_previous_bus) - 2][i]:
                            unique_time_bus_one_hour.insert(0,
                                                            time_all_previous_bus[len(time_all_previous_bus) - 2][
                                                                i])
                            break

                for i in range(len(unique_time_bus_one_hour)):  # ходим по времени прибытия уникального маршрута

                    if i != len(unique_time_bus_one_hour) - 1:  # защита от выхода
                        t1 = unique_time_bus_one_hour[i]  # берем первое время
                        ti = unique_time_bus_one_hour[i + 1]  # берем следущие вермя
                        t_dif = translation(
                            timedelta(hours=protection_time(ti)[0], minutes=protection_time(ti)[1])) - translation(
                            timedelta(hours=protection_time(t1)[0],
                                      minutes=protection_time(t1)[1]))  # находим разницу между приходами
                        # вычисляем абсолютную разницу между оптимальной разницой и разницой
                        dif = abs(optimal_time - t_dif)
                        check_prov_between_same_bus += dif  # получаем сумму отклонений, чем больше сумма тем хуже
                    # print(ti, t1, t_dif, dif)
                mas_second_check_prov_between_same_bus.append(check_prov_between_same_bus)
            else:
                mas_second_check_prov_between_same_bus.append(0)
        print("После оптимизации, абсолютное отклонение одного маршрута в ДУ =",
              mas_second_check_prov_between_same_bus, file=file)
        # print("После оптимизации, абсолютное отклонение одного маршрута в ДУ =",
        #       mas_second_check_prov_between_same_bus)
        # -------------------------------- Конец--------------------------------------------------------------------------

        # -------------------------------- Второй подсчет абсолютного отклонения для все маршрутов в ДУ--------------------------------------------------------------------------
        last_check_prov = 0
        flag_improve = True
        for time_bus_one_hour_x in range(len(time_bus_one_hour)):
            if flag_improve:
                # в данном ифе считаем с чек с предыдущим автобусом из прошлого часа
                t1 = time_last_previous_bus[len(time_last_previous_bus) - 2]
                # t1 = time_all_previous_bus[len(time_all_previous_bus) - 1][len(time_all_previous_bus[len(time_all_previous_bus) - 1]) - 1]
                ti = time_bus_one_hour[time_bus_one_hour_x]
                # находим разницу между приходами
                # из-за библиотеки приходится чекать и добавлять 24 часа
                # так когда отнимаем от 00:00 получается не правильное время
                if ti[:2] == "00":
                    t_dif = (translation(timedelta(hours=protection_time(ti)[0],
                                                   minutes=protection_time(ti)[1])) + 1440) - translation(
                        timedelta(hours=protection_time(t1)[0], minutes=protection_time(t1)[1]))
                else:
                    t_dif = translation(
                        timedelta(hours=protection_time(ti)[0], minutes=protection_time(ti)[1])) - translation(
                        timedelta(hours=protection_time(t1)[0], minutes=protection_time(t1)[1]))
                # вычисляем абсолютную разницу между оптимальной разницой и разницой
                dif = abs(minutes(time_pause) - t_dif)

                if flag_first_check:
                    dif = 0
                last_check_prov += dif  # получаем сумму отклонений, чем больше сумма тем хуже
                flag_improve = False
                # print("ti=", ti, 't1=', t1, 'dif=', dif)

            if time_bus_one_hour_x != len(time_bus_one_hour) - 1:  # защита от выхода
                t1 = time_bus_one_hour[time_bus_one_hour_x]  # берем первое время
                ti = time_bus_one_hour[time_bus_one_hour_x + 1]  # берем следущие вермя
                t_dif = translation(
                    timedelta(hours=protection_time(ti)[0], minutes=protection_time(ti)[1])) - translation(
                    timedelta(hours=protection_time(t1)[0],
                              minutes=protection_time(t1)[1]))  # находим разницу между приходами
                # вычисляем абсолютную разницу между оптимальной разницой и разницой
                dif = abs(minutes(time_pause) - t_dif)
                last_check_prov += dif  # получаем сумму отклонений, чем больше сумма тем хуже
                # print("ti=", ti, 't1=', t1, 'dif=', dif)
        print("После оптимизации, абсолютное отклонение по всем маршрутом в ДУ", last_check_prov, file=file)
        # print("После оптимизации, абсолютное отклонение по всем маршрутом в ДУ", last_check_prov)
        flag_first_check = False
        # -------------------------------- Конец--------------------------------------------------------------------------

        # print("")
        # print("-------------------------")
        # print(mas_sign)
        # print(mas_shift)
        # print(mas_direction)
        # print("-------------------------")
        # print("")
        # -------------------------------- Защита от ухудшения расписания--------------------------------------------------------------------------
        comeback = False
        for i in range(len(mas_second_check_prov_between_same_bus)):  # ходим по абсолютным величнам
            # если до оптимизации было лучше,то
            if mas_second_check_prov_between_same_bus[i] > mas_first_check_prov_between_same_bus[i] \
                    or last_check_prov > first_check_prov:
                comeback = True
                break
        #  для проверки изменений
        all_check_prov[0] += first_check_prov
        if comeback:
            all_check_prov[1] += first_check_prov
        elif not comeback:
            all_check_prov[1] += last_check_prov
        if comeback:  # то выполняем откат
            # выполняем откад индексов автобусов
            name_previous_bus[len(time_all_previous_bus) - 1] = temp_index
            for time_bus_one_hour_x in range(len(time_bus_one_hour)):  # ходим по времени прибытия

                num_bus = index_time_bus_one_hour[time_bus_one_hour_x]  # кусок кода был уже применен в оптимизации
                num_tab = name.index(num_bus)  # Ищем имя первого автобуса в массиве name
                num_bus_in_dy = dy[1].index(num_bus)

                # Бегаем вертикально по строкам времени приходом автобуса на базовую остановку
                for y in range(2, len_row[num_tab] + 1):
                    # Если время совпало с временем из таблицы(таким образом находим строчку нужного нам рейса)
                    if str(old_time_bus_one_hour[time_bus_one_hour_x]) == str(
                            schedule[num_tab][y][dy[2][num_bus_in_dy]]):

                        # if index_time_bus_one_hour[time_bus_one_hour_x] == check_name:  # название маршрута совпала
                        # то исправляем time_bus
                        # откатываем время прихода на БО и Якорное время
                        time_bus_one_hour[time_bus_one_hour_x] = old_time_bus_one_hour[time_bus_one_hour_x]
                        anchor_time_bus_one_hour[time_bus_one_hour_x] = anchor_old_time_bus_one_hour[
                            time_bus_one_hour_x]

                        if mas_live_bus[time_bus_one_hour_x] != 0:
                            # если автобус находился в прямом направление
                            if mas_direction[time_bus_one_hour_x] == "Ahead":
                                # идем от начала таблицы до конца прямого направления
                                for shift in range(2, col_reverse[num_tab]):
                                    if schedule[num_tab][y][shift] is not None:
                                        if schedule_with_anchor[num_tab][y][shift] != 1:
                                            if type(schedule[num_tab][y][shift]) != str:
                                                schedule[num_tab][y][shift] = str(schedule[num_tab][y][shift])
                                            temp = schedule[num_tab][y][shift]
                                            # print("schedule до сдвига", schedule[num_tab][y][shift])
                                            temp = timedelta(hours=protection_time(temp)[0],
                                                             minutes=protection_time(temp)[1])
                                            # делаем знак на оборот, т.к. мы возвращаем время в обратное состояние
                                            if mas_sign[time_bus_one_hour_x] == "-":
                                                sign = "+"
                                            else:
                                                sign = "-"
                                            schedule[num_tab][y][shift] = str(
                                                arithmetic(temp, timedelta(minutes=mas_shift[time_bus_one_hour_x]),
                                                           sign))
                                            # print("schedule после сдвига", schedule[num_tab][y][shift],"\n")

                            # если автобус находился в обратном направление
                            elif mas_direction[time_bus_one_hour_x] == "Back":
                                # идем от конца таблице до начала обратного направления
                                for shift in range(len_col[num_tab] - 1, col_reverse[num_tab] - 1, -1):
                                    if schedule[num_tab][y][shift] is not None:
                                        if schedule_with_anchor[num_tab][y][shift] != 1:
                                            temp = schedule[num_tab][y][shift]
                                            # print("schedule до сдвига", schedule[num_tab][y][shift])
                                            temp = timedelta(hours=protection_time(temp)[0],
                                                             minutes=protection_time(temp)[1])
                                            if mas_sign[time_bus_one_hour_x] == "-":
                                                sign = "+"
                                            else:
                                                sign = "-"
                                            schedule[num_tab][y][shift] = str(
                                                arithmetic(temp, timedelta(minutes=mas_shift[time_bus_one_hour_x]),
                                                           sign))
                                            # print("schedule после сдвига", schedule[num_tab][y][shift])

            # -------------------------------- Конец--------------------------------------------------------------------------
            # print(time_all_previous_bus[hour])

            # изменяем и массив который используем для обратки время для подсчета перерывов
            time_all_previous_bus[len(time_all_previous_bus) - 1] = time_bus_one_hour
            time_last_previous_bus[len(time_last_previous_bus) - 1] = time_bus_one_hour[len(time_bus_one_hour) - 1]

            print("После проверки:", time_bus_one_hour, file=file)
            # print("После проверки:", time_bus_one_hour)

name_new_folder = f'./Расписание других городов/{name_city}/New schedule/'  # название нового каталога
if not os.path.isdir(name_new_folder):  # если католога нет
    os.mkdir(name_new_folder)  # то создаем

for NAME in range(len(name)):  # ходим по всем маршрутам
    for row in range(2, len_row[NAME] + 1):  # бегаем по строкам в таблице данного маршрутаа
        if schedule[NAME][row][0] != schedule[NAME][row][2]:
            schedule[NAME][row][0] = schedule[NAME][row][2]

for NAME in range(len(name)):  # ходим по всем маршрутам
    # print()
    # print(name[NAME],'-------------------------------------------------------------')
    workbook = xlsxwriter.Workbook(str(os.getcwd()) + "\\" + str(name_new_folder) + "\\" + str(
        name[NAME]) + '.xlsx')  # указываем маршрут куда будем сохрянять
    worksheet = workbook.add_worksheet()
    #  записываем клетку с началам обратного напрваления, а если его нет, то просто записываем прямое
    if col_reverse[NAME] == len_col[NAME]:
        worksheet.write(0, 3, "Прямое направление")
    else:
        worksheet.write(0, col_reverse[NAME], "Обратное направление")
    for row in range(len_row[NAME] + 1):  # бегаем по строкам в таблице данного маршрута
        # print(schedule[NAME][row])
        for col in range(len_col[NAME]):  # бегаем по столбцам в таблице данного маршрута
            if schedule[NAME][row][col] is None:  # если нон то ничего не записываем
                continue
            if type(schedule[NAME][row][col]) is not str():  # если не строка переделываем в строку
                schedule[NAME][row][col] = str(schedule[NAME][row][col])
                # print(schedule[NAME][row][col])
            worksheet.write(row, col, schedule[NAME][row][col])
    workbook.close()
print(all_check_prov[0], all_check_prov[1])


# workbook = on.load_workbook(f'./Расписание других городов/{name_city}/result.xlsx')
# worksheet = workbook["1"]
# worksheet.cell(row=int(options[4]), column=5, value=int(quantity_dy))
# worksheet.cell(row=int(options[4]), column=6, value=options[1])
# worksheet.cell(row=int(options[4]), column=7, value=options[2])
# worksheet.cell(row=int(options[4]), column=8, value=options[3])
# worksheet.cell(row=int(options[4]), column=9, value=all_check_prov[0])
# print("options[4]=",options[4])
# worksheet.cell(row=int(options[4]), column=10, value=all_check_prov[1])
# # Сохраняем изменения
# workbook.save(f'./Расписание других городов/{name_city}/result.xlsx')
# res = subprocess.run(['python', 'improvements.py', options[4]], stdout=subprocess.PIPE, stderr=subprocess.PIPE,
#                      text=True,
#                      encoding='utf-8')
# v = res.stdout
# print(v)